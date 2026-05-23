from __future__ import annotations

import csv
import json
import os
import secrets
import uuid
import warnings
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path

from flask import Blueprint, Response, current_app, jsonify, redirect, request

import db
from audit import audit
from auth import require_role
from validators import ValidationError, body, integer, raw_text


api_bp = Blueprint("features_api", __name__, url_prefix="/api")
seo_bp = Blueprint("seo", __name__)
_settings_cache: dict = {"ts": 0, "data": None}


def ist_tz():
    try:
        import zoneinfo
        return zoneinfo.ZoneInfo("Asia/Kolkata")
    except Exception:
        return timezone(timedelta(hours=5, minutes=30), "IST")


def _active_rules(conn) -> list[dict]:
    now = datetime.now(ist_tz())
    current_time = now.strftime("%H:%M")
    rows = conn.execute(
        "SELECT * FROM pricing_rules WHERE active = true AND start_time <= ? AND end_time > ?",
        (current_time, current_time),
    ).fetchall()
    result = []
    for row in rows:
        days = db.decode_json(row["days_of_week"], [0, 1, 2, 3, 4, 5, 6])
        if now.weekday() in days:
            result.append(dict(row))
    return result


@api_bp.post("/feedback")
def submit_feedback():
    data = body()
    order_id = integer(data.get("order_id"), "order_id", 1)
    rating = integer(data.get("rating"), "rating", 1, 5)
    comment = raw_text(data.get("comment", ""), "comment", 500, required=False, allow_empty=True)

    def write(conn):
        conn.execute(
            "INSERT INTO feedback (order_id, rating, comment, created_at) VALUES (?, ?, ?, ?) ON CONFLICT (order_id) DO NOTHING",
            (order_id, rating, comment, db.utc_now()),
        )

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"status": "ok"})


@api_bp.post("/coupons/apply")
def apply_coupon():
    data = body()
    code = str(data.get("code", "")).strip().upper()
    subtotal = integer(data.get("subtotal"), "subtotal", 0)

    if not code:
        return jsonify({"success": False, "message": "Heritage code required."}), 400

    def check(conn):
        # Using raw SQL for compatibility with existing features.py patterns
        coupon = conn.execute(
            "SELECT * FROM campaigns WHERE code = ? AND active = true",
            (code,)
        ).fetchone()

        if not coupon:
            return None, "This heritage code is not recognized."

        now = db.utc_now()
        # Handle date strings from raw SQL
        start_date = datetime.fromisoformat(coupon["start_date"].replace('Z', '+00:00')) if isinstance(coupon["start_date"], str) else coupon["start_date"]
        end_date = datetime.fromisoformat(coupon["end_date"].replace('Z', '+00:00')) if isinstance(coupon["end_date"], str) else coupon["end_date"]

        if now < start_date:
            return None, "This campaign has not yet commenced."
        if now > end_date:
            return None, "This campaign has concluded its journey."
        
        if coupon["usage_limit"] and coupon["usage_count"] >= coupon["usage_limit"]:
            return None, "This code has reached its maximum legacy."
            
        if subtotal < coupon["min_order_value"]:
            return None, f"Minimum investment of ₹{coupon['min_order_value'] / 100:.2f} required."

        return dict(coupon), None

    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        coupon, error = check(conn)
    
    if error:
        return jsonify({"success": False, "message": error}), 400

    discount = 0
    if coupon["type"] == "percentage":
        discount = int(subtotal * (coupon["value"] / 100))
    else:
        discount = coupon["value"]

    return jsonify({
        "success": True,
        "discount": discount,
        "code": code,
        "title": coupon["title"]
    })


@api_bp.post("/vouchers/check")
def check_voucher():
    data = body()
    code = str(data.get("code", "")).strip().upper()

    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        voucher = conn.execute(
            "SELECT * FROM vouchers WHERE code = ?",
            (code,)
        ).fetchone()

    if not voucher:
        return jsonify({"success": False, "message": "Voucher not found."}), 404
    
    # Handle date
    expires_at = datetime.fromisoformat(voucher["expires_at"].replace('Z', '+00:00')) if isinstance(voucher["expires_at"], str) else voucher["expires_at"]
    if db.utc_now() > expires_at:
        return jsonify({"success": False, "message": "Voucher expired."}), 400
    
    if voucher["current_value"] <= 0:
        return jsonify({"success": False, "message": "Voucher has no remaining balance."}), 400

    return jsonify({
        "success": True,
        "balance": voucher["current_value"],
        "code": code
    })


@api_bp.get("/settings/public")
def public_settings():
    import time

    now = time.monotonic()
    if _settings_cache["data"] and now - _settings_cache["ts"] < 3600:
        return jsonify(_settings_cache["data"])
    keys = ["google_rating", "google_review_count", "google_place_id", "google_review_url", "restaurant_name", "restaurant_address", "restaurant_phone"]
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        rows = conn.execute(f"SELECT key, value FROM site_settings WHERE key IN ({','.join('?' * len(keys))})", keys).fetchall()
    data = {r["key"]: r["value"] for r in rows}
    _settings_cache.update({"ts": now, "data": data})
    return jsonify(data)


@api_bp.patch("/admin/settings")
@require_role("admin")
def patch_settings():
    data = body()
    allowed = {"google_rating", "google_review_count", "google_place_id", "google_review_url", "restaurant_name", "restaurant_address", "restaurant_phone", "restaurant_hours", "pickup_max_per_slot"}

    def write(conn):
        clean = {}
        for key, value in data.items():
            if key not in allowed:
                continue
            clean[key] = str(value)
            conn.execute(
                "INSERT INTO site_settings (key, value, updated_at) VALUES (?, ?, ?) ON CONFLICT (key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
            (key, str(value), db.utc_now()),
            )
        audit(conn, "settings.public_update", "site_settings", "all", clean)

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    _settings_cache["ts"] = 0
    return jsonify({"status": "ok"})


@api_bp.get("/pricing/active")
def active_pricing():
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        rules = _active_rules(conn)
    return jsonify({"rules": rules})


@api_bp.post("/admin/pricing-rules")
@require_role("admin")
def create_pricing_rule():
    data = body()
    payload = _pricing_payload(data)

    def write(conn):
        cursor = conn.execute(
            """
            INSERT INTO pricing_rules (name, days_of_week, start_time, end_time, discount_type, discount_value,
                                       applies_to, applies_to_ids, active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*payload, db.utc_now()),
        )
        audit(conn, "pricing.create", "pricing_rule", cursor.lastrowid)
        return {"id": cursor.lastrowid}

    return jsonify(db.atomic_write(write, current_app.config["DATABASE_URL"])), 201


@api_bp.patch("/admin/pricing-rules/<int:rule_id>")
@require_role("admin")
def update_pricing_rule(rule_id: int):
    data = body()
    payload = _pricing_payload(data)

    def write(conn):
        row = conn.execute("SELECT id FROM pricing_rules WHERE id = ?", (rule_id,)).fetchone()
        if not row:
            raise ValidationError("Pricing rule not found", "rule_id", 404)
        conn.execute(
            """
            UPDATE pricing_rules
            SET name = ?, days_of_week = ?, start_time = ?, end_time = ?, discount_type = ?, discount_value = ?,
                applies_to = ?, applies_to_ids = ?, active = ?
            WHERE id = ?
            """,
            (*payload, rule_id),
        )
        audit(conn, "pricing.update", "pricing_rule", rule_id)

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"status": "ok"})


@api_bp.delete("/admin/pricing-rules/<int:rule_id>")
@require_role("admin")
def delete_pricing_rule(rule_id: int):
    def write(conn):
        conn.execute("DELETE FROM pricing_rules WHERE id = ?", (rule_id,))
        audit(conn, "pricing.delete", "pricing_rule", rule_id)

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"status": "ok"})


def _pricing_payload(data: dict) -> tuple:
    discount_type = raw_text(data.get("discount_type"), "discount_type", 20)
    if discount_type not in {"percent", "fixed"}:
        raise ValidationError("Invalid discount_type", "discount_type")
    days = data.get("days_of_week", [0, 1, 2, 3, 4, 5, 6])
    ids = data.get("applies_to_ids", [])
    if not isinstance(days, list) or not all(isinstance(x, int) and 0 <= x <= 6 for x in days):
        raise ValidationError("days_of_week must be weekday integers", "days_of_week")
    if not isinstance(ids, list):
        raise ValidationError("applies_to_ids must be a list", "applies_to_ids")
    return (
        raw_text(data.get("name"), "name", 120),
        db.encode_json(days),
        raw_text(data.get("start_time"), "start_time", 5),
        raw_text(data.get("end_time"), "end_time", 5),
        discount_type,
        float(data.get("discount_value", 0)),
        raw_text(data.get("applies_to", "all"), "applies_to", 40, required=False) or "all",
        db.encode_json(ids),
        1 if bool(data.get("active", True)) else 0,
    )


@api_bp.post("/ratings")
@require_role("customer")
def submit_ratings():
    data = body()
    order_id = integer(data.get("order_id"), "order_id", 1)
    ratings_data = data.get("ratings", [])
    if not isinstance(ratings_data, list) or not ratings_data:
        raise ValidationError("ratings are required", "ratings")
    user = getattr(__import__("flask").g, "current_user", None)

    def write(conn):
        order = conn.execute("SELECT * FROM orders WHERE id = ? AND (user_id = ? OR ? IN ('staff','admin'))", (order_id, user["id"], user["role"])).fetchone()
        if not order:
            raise ValidationError("Order not found", "order_id", 404)
        for item in ratings_data[:20]:
            if not isinstance(item, dict):
                continue
            mid = integer(item.get("menu_item_id"), "menu_item_id", 1, required=False)
            rat = integer(item.get("rating"), "rating", 1, 5, required=False)
            if mid and rat:
                conn.execute(
                    "INSERT INTO dish_ratings (order_id, menu_item_id, user_id, rating, created_at) VALUES (?, ?, ?, ?, ?) ON CONFLICT (order_id, menu_item_id) DO NOTHING",
            (order_id, mid, user["id"], rat, db.utc_now()),
                )
    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"status": "ok"})


@api_bp.get("/qr-gen")
def generate_branded_qr():
    import qrcode
    from io import BytesIO

    data = request.args.get("data", "https://jayadhaba.com")
    
    # Branded QR logic
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="#1A0F0A", back_color="#FAF9F6")
    
    buf = BytesIO()
    img.save(buf)
    buf.seek(0)
    
    return Response(buf.read(), mimetype="image/png")


@api_bp.post("/ratings/<int:rating_id>/photo")
@require_role("customer")
def upload_rating_photo(rating_id: int):
    from PIL import Image

    f = request.files.get("photo")
    if not f:
        raise ValidationError("photo is required", "photo")
    if f.mimetype not in ("image/jpeg", "image/png", "image/webp"):
        raise ValidationError("JPG, PNG, or WebP only", "photo")
    data = f.read()
    if len(data) > 5 * 1024 * 1024:
        raise ValidationError("Max 5MB", "photo", 413)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            img = Image.open(BytesIO(data))
            img.verify()
            img = Image.open(BytesIO(data)).convert("RGB")
    except Exception:
        raise ValidationError("Invalid image content", "photo")
    img.thumbnail((800, 800), Image.LANCZOS)
    fname = f"community_{rating_id}_{uuid.uuid4().hex[:8]}.webp"
    path = Path(current_app.config["UPLOAD_FOLDER"]) / "community" / fname
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path, format="WEBP", quality=80)
    photo_url = f"/uploads/community/{fname}"

    def write(conn):
        conn.execute("UPDATE dish_ratings SET photo_url = ? WHERE id = ?", (photo_url, rating_id))

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"photo_url": photo_url})


@api_bp.get("/admin/ratings/pending")
@require_role("staff")
def pending_ratings():
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        rows = conn.execute(
            """
            SELECT dr.*, mi.name AS dish_name FROM dish_ratings dr
            JOIN menu_items mi ON mi.id = dr.menu_item_id
            WHERE dr.photo_url IS NOT NULL AND dr.photo_approved = 0
            ORDER BY dr.created_at DESC LIMIT 50
            """
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@api_bp.patch("/admin/ratings/<int:rating_id>")
@require_role("staff")
def moderate_rating(rating_id: int):
    approved = bool(body().get("approved"))

    def write(conn):
        conn.execute("UPDATE dish_ratings SET photo_approved = ? WHERE id = ?", (1 if approved else 0, rating_id))
        audit(conn, "rating.moderate", "dish_rating", rating_id, {"approved": approved})

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"status": "ok"})


def generate_referral_code() -> str:
    return secrets.token_urlsafe(6).upper()[:8]


def generate_reward_card_image(reward_card_id: str, reward_type: str) -> str:
    import qrcode
    from PIL import Image, ImageDraw, ImageFont

    w, h = 800, 450
    img = Image.new("RGB", (w, h), (20, 15, 8))
    draw = ImageDraw.Draw(img)
    draw.rectangle([10, 10, w - 10, h - 10], outline=(212, 175, 55), width=3)
    try:
        font_lg = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 48)
        font_md = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 28)
        font_sm = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 20)
    except Exception:
        font_lg = font_md = font_sm = ImageFont.load_default()
    draw.text((40, 40), "JAYA DHABA", fill=(212, 175, 55), font=font_lg)
    draw.text((40, 110), "1 Free Cool Drink", fill=(255, 255, 255), font=font_md)
    draw.text((40, 160), "Present this card to staff", fill=(180, 180, 180), font=font_sm)
    draw.text((40, 190), f"Card: {reward_card_id[:12]}", fill=(150, 150, 150), font=font_sm)
    qr = qrcode.make(reward_card_id).resize((180, 180))
    img.paste(qr, (w - 220, h - 220))
    path = Path(current_app.config["UPLOAD_FOLDER"]) / "rewards" / f"reward_{reward_card_id}.png"
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path, format="PNG")
    return f"/uploads/rewards/reward_{reward_card_id}.png"


def _card_data(row) -> dict | None:
    if not row["reward_card_id"]:
        return None
    return {"card_id": row["reward_card_id"], "reward_type": row["reward_type"], "claimed": bool(row["reward_claimed"]), "redeemed": bool(row["reward_redeemed"]), "image_url": f"/uploads/rewards/reward_{row['reward_card_id']}.png"}


@api_bp.get("/referral/my-code")
@require_role("customer")
def my_referral_code():
    user = getattr(__import__("flask").g, "current_user", None)
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        row = conn.execute("SELECT * FROM referrals WHERE referrer_user_id = ?", (user["id"],)).fetchone()
        if row:
            return jsonify({"code": row["referral_code"], "link": f"{current_app.config['DOMAIN']}/r/{row['referral_code']}", "reward_card": _card_data(row)})
    code = generate_referral_code()

    def write(conn):
        conn.execute("INSERT INTO referrals (referrer_user_id, referral_code, created_at) VALUES (?, ?, ?)", (user["id"], code, db.utc_now()))

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"code": code, "link": f"{current_app.config['DOMAIN']}/r/{code}", "reward_card": None})


@seo_bp.get("/r/<code>")
def referral_redirect(code):
    resp = redirect(f"/?ref={code}")
    resp.set_cookie("ref_code", code, max_age=86400, httponly=True, samesite="Lax")
    return resp


@api_bp.post("/admin/referral/verify")
@require_role("staff")
def verify_reward():
    card_id = raw_text(body().get("reward_card_id", ""), "reward_card_id", 120, required=False, allow_empty=True)
    if not card_id:
        raise ValidationError("reward_card_id is required", "reward_card_id")
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        row = conn.execute("SELECT * FROM referrals WHERE reward_card_id = ?", (card_id,)).fetchone()
    if not row:
        return jsonify({"valid": False, "reason": "Card not found"})
    if row["reward_redeemed"]:
        return jsonify({"valid": False, "reason": "Already redeemed", "redeemed_at": row["redeemed_at"]})
    return jsonify({"valid": True, "reward_type": row["reward_type"]})


@api_bp.post("/admin/referral/redeem")
@require_role("staff")
def redeem_reward():
    card_id = raw_text(body().get("reward_card_id", ""), "reward_card_id", 120)

    def write(conn):
        row = conn.execute("SELECT * FROM referrals WHERE reward_card_id = ?", (card_id,)).fetchone()
        if not row:
            raise ValidationError("Card not found", "reward_card_id", 404)
        if row["reward_redeemed"]:
            raise ValidationError("Already redeemed", "reward_card_id", 409)
        conn.execute("UPDATE referrals SET reward_redeemed = 1, redeemed_at = ?, redeemed_by_admin = ? WHERE reward_card_id = ?", (db.utc_now(), getattr(__import__("flask").g, "current_user", {}).get("id"), card_id))
        audit(conn, "referral.redeem", "referral", card_id)

    db.atomic_write(write, current_app.config["DATABASE_URL"])
    return jsonify({"status": "ok"})


@api_bp.get("/admin/briefing")
@require_role("admin")
def daily_briefing():
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        cached = conn.execute("SELECT value FROM site_settings WHERE key='briefing_cache'").fetchone()
        cached_at = conn.execute("SELECT value FROM site_settings WHERE key='briefing_cached_at'").fetchone()
        now = datetime.now(ist_tz())
        cache_until = now.replace(hour=6, minute=0, second=0, microsecond=0)
        if now.hour >= 6:
            cache_until = cache_until + __import__("datetime").timedelta(days=1)
        if cached and cached["value"] and cached_at and cached_at["value"] > now.isoformat():
            return jsonify(json.loads(cached["value"]))
        today = now.date().isoformat()
        y_rev = conn.execute("SELECT COALESCE(SUM(total),0) AS rev, COUNT(*) AS cnt FROM orders WHERE status!='cancelled' AND date(created_at)=date('now','-1 day')").fetchone()
        db_rev = conn.execute("SELECT COALESCE(SUM(total),0) AS rev FROM orders WHERE status!='cancelled' AND date(created_at)=date('now','-2 days')").fetchone()["rev"]
        top = conn.execute(
            """
            SELECT mi.name, SUM(oi.qty) AS total_qty FROM order_items oi
            JOIN menu_items mi ON mi.id=oi.menu_item_id JOIN orders o ON o.id=oi.order_id
            WHERE date(o.created_at)=date('now','-1 day') AND o.status!='cancelled'
            GROUP BY oi.menu_item_id ORDER BY total_qty DESC LIMIT 3
            """
        ).fetchall()
        from db import engine as _engine
        if _engine.dialect.name == "sqlite":
            _hr_expr = "strftime('%H', created_at)"
            _yesterday = "date(created_at) = date('now', '-1 day')"
        else:
            _hr_expr = "to_char(created_at AT TIME ZONE 'Asia/Kolkata', 'HH24')"
            _yesterday = "created_at::date = (NOW() AT TIME ZONE 'Asia/Kolkata')::date - INTERVAL '1 day'"
        busiest = conn.execute(
            f"SELECT {_hr_expr} AS hr, COUNT(*) AS cnt FROM orders WHERE {_yesterday} GROUP BY hr ORDER BY cnt DESC LIMIT 1"
        ).fetchone()
        stats = {
            "yesterday_revenue": round(y_rev["rev"], 2),
            "yesterday_orders": y_rev["cnt"],
            "revenue_change_pct": round(((y_rev["rev"] - db_rev) / db_rev * 100) if db_rev > 0 else 0, 1),
            "top_dishes": [{"name": r["name"], "qty": r["total_qty"]} for r in top],
            "busiest_hour": f"{busiest['hr']}:00" if busiest else None,
            "reservations_today": conn.execute("SELECT COUNT(*) FROM reservations WHERE date(reserved_at)=? AND status='confirmed'", (today,)).fetchone()[0],
            "pending_photos": conn.execute("SELECT COUNT(*) FROM dish_ratings WHERE photo_url IS NOT NULL AND photo_approved=0").fetchone()[0],
            "pickup_today": conn.execute("SELECT COUNT(*) FROM orders WHERE order_type='pickup' AND date(created_at)=?", (today,)).fetchone()[0],
        }
        if stats["yesterday_orders"] == 0:
            stats["message"] = "First day! No data yet. You've got this."
        conn.execute(
    "INSERT INTO site_settings (key, value, updated_at) VALUES ('briefing_cache', ?, ?) ON CONFLICT (key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
    (json.dumps(stats), db.utc_now())
)
        conn.execute(
    "INSERT INTO site_settings (key, value, updated_at) VALUES ('briefing_cached_at', ?, ?) ON CONFLICT (key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at",
    (cache_until.isoformat(), db.utc_now())
)
        conn.commit()
    return jsonify(stats)


@api_bp.get("/admin/customers")
@require_role("admin")
def customers():
    q = request.args.get("q", "").strip()[:100]
    page = max(1, int(request.args.get("page", 1)))
    params = []
    where = "WHERE 1=1"
    if q:
        where += " AND (u.email LIKE ? OR u.phone LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        rows = conn.execute(
            f"""
            SELECT u.id,u.email,u.phone,u.created_at,u.loyalty_points,
                   COUNT(DISTINCT o.id) AS visit_count, COALESCE(SUM(o.total),0) AS total_spend,
                   COALESCE(AVG(o.total),0) AS avg_order, MAX(o.created_at) AS last_visit
            FROM users u LEFT JOIN orders o ON o.user_id=u.id AND o.status!='cancelled'
            {where}
            GROUP BY u.id ORDER BY last_visit DESC LIMIT ? OFFSET ?
            """,
            params + [20, (page - 1) * 20],
        ).fetchall()
    return jsonify({"customers": [dict(r) for r in rows], "page": page})


@api_bp.get("/admin/customers/<int:user_id>")
@require_role("admin")
def customer_profile(user_id: int):
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        user = conn.execute("SELECT id,email,phone,loyalty_points,created_at FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            raise ValidationError("Customer not found", "user_id", 404)
        orders = conn.execute("SELECT id,total,status,created_at FROM orders WHERE user_id=? ORDER BY created_at DESC LIMIT 50", (user_id,)).fetchall()
        fav = conn.execute(
            """
            SELECT mi.name, COUNT(*) AS cnt FROM order_items oi
            JOIN orders o ON o.id=oi.order_id JOIN menu_items mi ON mi.id=oi.menu_item_id
            WHERE o.user_id=? GROUP BY oi.menu_item_id ORDER BY cnt DESC LIMIT 1
            """,
            (user_id,),
        ).fetchone()
        feedback = conn.execute("SELECT AVG(f.rating) AS avg_rating, COUNT(*) AS cnt FROM feedback f JOIN orders o ON o.id=f.order_id WHERE o.user_id=?", (user_id,)).fetchone()
    return jsonify({"user": dict(user), "orders": [dict(r) for r in orders], "favorite_dish": dict(fav) if fav else None, "feedback": {"avg_rating": round(feedback["avg_rating"] or 0, 1), "count": feedback["cnt"]}})


@api_bp.get("/admin/export")
@require_role("admin", missing_status=403)
def export_report():
    report_type = request.args.get("type", "orders")
    fmt = request.args.get("format", "csv")
    date_from = request.args.get("from", "2000-01-01")
    date_to = request.args.get("to", "2099-12-31")
    if report_type not in {"orders", "revenue", "customers", "feedback"} or fmt not in {"csv", "excel", "pdf"}:
        raise ValidationError("Invalid export request")
    headers, data = _export_data(report_type, date_from, date_to)
    filename = f"jaya-dhaba-{report_type}-{date_from}-to-{date_to}"
    if fmt == "excel":
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.title()
        ws.append(headers)
        for row in data:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"})
    if fmt == "pdf":
        try:
            from fpdf import FPDF

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, f"Jaya Dhaba - {report_type.title()} Report", ln=True)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 6, f"Period: {date_from} to {date_to}", ln=True)
            pdf.set_font("Helvetica", "B", 8)
            col_w = max(20, 190 // len(headers))
            for h in headers:
                pdf.cell(col_w, 6, str(h)[:24], border=1)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            for row in data:
                for cell in row:
                    pdf.cell(col_w, 5, str(cell)[:30], border=1)
                pdf.ln()
            return Response(bytes(pdf.output(dest="S")), mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}.pdf"})
        except Exception:
            fmt = "csv"
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(data)
    return Response(buf.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}.csv"})


def _export_data(report_type: str, date_from: str, date_to: str):
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        if report_type == "customers":
            rows = conn.execute("SELECT id,email,phone,loyalty_points,created_at FROM users ORDER BY created_at DESC LIMIT 5000").fetchall()
            return ["ID", "Email", "Phone", "Points", "Created"], [[r["id"], r["email"], r["phone"], r["loyalty_points"], r["created_at"]] for r in rows]
        if report_type == "feedback":
            rows = conn.execute("SELECT f.order_id,f.rating,f.comment,f.created_at FROM feedback f WHERE date(f.created_at) BETWEEN ? AND ? ORDER BY f.created_at DESC LIMIT 5000", (date_from, date_to)).fetchall()
            return ["Order", "Rating", "Comment", "Created"], [[r["order_id"], r["rating"], r["comment"], r["created_at"]] for r in rows]
        if report_type == "revenue":
            rows = conn.execute("SELECT date(created_at) AS day, COALESCE(SUM(total),0) AS revenue, COUNT(*) AS orders FROM orders WHERE date(created_at) BETWEEN ? AND ? AND status!='cancelled' GROUP BY date(created_at) ORDER BY day DESC LIMIT 5000", (date_from, date_to)).fetchall()
            return ["Day", "Revenue", "Orders"], [[r["day"], r["revenue"], r["orders"]] for r in rows]
        rows = conn.execute(
            """
            SELECT o.id,o.status,o.total,o.order_type,o.created_at,u.email
            FROM orders o LEFT JOIN users u ON u.id=o.user_id
            WHERE date(o.created_at) BETWEEN ? AND ?
            ORDER BY o.created_at DESC LIMIT 5000
            """,
            (date_from, date_to),
        ).fetchall()
        return ["ID", "Status", "Total", "Type", "Created", "Customer"], [[r["id"], r["status"], r["total"], r["order_type"], r["created_at"], r["email"] or "Guest"] for r in rows]


@seo_bp.get("/sitemap.xml")
def sitemap():
    domain = os.environ.get("DOMAIN", "https://yourdomain.com").rstrip("/")
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        cats = conn.execute("SELECT id FROM menu_categories WHERE active=1").fetchall()
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for page in ["", "/menu", "/reservations", "/track"]:
        lines.append(f"<url><loc>{domain}{page}</loc><changefreq>weekly</changefreq></url>")
    for cat in cats:
        lines.append(f"<url><loc>{domain}/menu?category={cat['id']}</loc><changefreq>daily</changefreq></url>")
    lines.append("</urlset>")
    return Response("\n".join(lines), mimetype="application/xml")


@api_bp.get("/seo/structured-data")
def structured_data():
    domain = os.environ.get("DOMAIN", "https://yourdomain.com").rstrip("/")
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        settings = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM site_settings").fetchall()}
    ld = {"@context": "https://schema.org", "@type": "Restaurant", "name": settings.get("restaurant_name", "Jaya Dhaba"), "address": {"@type": "PostalAddress", "streetAddress": settings.get("restaurant_address", "")}, "telephone": settings.get("restaurant_phone", ""), "url": domain, "menu": f"{domain}/menu"}
    if settings.get("google_rating", "0") != "0":
        ld["aggregateRating"] = {"@type": "AggregateRating", "ratingValue": settings.get("google_rating", "0"), "reviewCount": settings.get("google_review_count", "0")}
    return jsonify(ld)


@api_bp.get("/og-image")
def og_image():
    from PIL import Image, ImageDraw, ImageFont

    img = Image.new("RGB", (1200, 630), color=(30, 20, 10))
    draw = ImageDraw.Draw(img)
    with db.connect(current_app.config["DATABASE_URL"]) as conn:
        row = conn.execute("SELECT value FROM site_settings WHERE key='restaurant_name'").fetchone()
    name = row["value"] if row else "Jaya Dhaba"
    try:
        font_lg = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 80)
        font_sm = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 40)
    except Exception:
        font_lg = font_sm = ImageFont.load_default()
    draw.text((80, 200), name, fill=(212, 175, 55), font=font_lg)
    draw.text((80, 320), "Authentic Indian Cuisine - Hyderabad", fill=(200, 200, 200), font=font_sm)
    buf = BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return Response(buf.getvalue(), mimetype="image/png", headers={"Cache-Control": "public,max-age=86400"})
